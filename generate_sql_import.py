#!/usr/bin/env python3
"""
Script para converter dados do arquivo ODS em INSERTs SQL para MySQL
Requer: pip install openpyxl odfpy
"""

import openpyxl
from datetime import datetime
import re
import sys
import os

def load_ods_or_xlsx(file_path):
    """Carrega arquivo ODS ou XLSX"""
    if file_path.lower().endswith('.ods'):
        try:
            from odf import opendocument, table
            ods_doc = opendocument.load(file_path)
            sheets = ods_doc.spreadsheet.getElementsByType(table.Table)
            if not sheets:
                raise Exception("Nenhuma planilha encontrada")
            return sheets[0]
        except ImportError:
            print("Erro: biblioteca odf não instalada. Execute: pip install odfpy")
            sys.exit(1)
    else:
        return openpyxl.load_workbook(file_path).active

class OdsToSqlGenerator:
    def __init__(self, ods_file):
        self.ods_file = ods_file
        self.sql_statements = []
        self.ativo_map = {}
        self.tipo_map = {}
        self.secretaria_map = {}
        self.servidor_map = {}
        self.funcao_map = {}
        self.next_ativo_id = 1
        self.next_tipo_id = 1
        self.next_secretaria_id = 1
        self.next_servidor_id = 1
        self.next_funcao_id = 1
        self.next_equipe_id = 1
        self.next_ata_id = 1

    def load_workbook(self):
        """Carrega o arquivo Excel/ODS"""
        try:
            if self.ods_file.lower().endswith('.ods'):
                from odf import opendocument, table, text
                
                ods_doc = opendocument.load(self.ods_file)
                sheets = ods_doc.spreadsheet.getElementsByType(table.Table)
                if not sheets:
                    raise Exception("Nenhuma planilha encontrada")
                
                return sheets[0], 'ods'
            else:
                return openpyxl.load_workbook(self.ods_file).active, 'xlsx'
        except ImportError as e:
            print(f"Erro: Biblioteca faltante - {e}")
            print("Instale com: pip install odfpy")
            sys.exit(1)
        except Exception as e:
            print(f"Erro ao carregar arquivo: {e}")
            sys.exit(1)

    def escape_sql_string(self, value):
        """Escapa string para SQL"""
        if value is None:
            return "NULL"
        value_str = str(value).replace("'", "''")
        return f"'{value_str}'"

    def format_date(self, value):
        """Formata data para SQL"""
        if value is None:
            return "NULL"
        try:
            if isinstance(value, str):
                # Tenta parsear DD/MM/YYYY
                if '/' in value:
                    parts = value.split('/')
                    if len(parts) == 3:
                        return f"'{parts[2]}-{parts[1]}-{parts[0]}'"
            elif hasattr(value, 'date'):
                return f"'{value.date()}'"
            elif hasattr(value, 'strftime'):
                return f"'{value.strftime('%Y-%m-%d')}'"
        except:
            pass
        return "NULL"

    def parse_gestores_fiscais(self, text):
        """Parseia gestores e fiscais do texto"""
        if not text:
            return [], []
        
        gestores = []
        fiscais = []
        
        linhas = text.split('\n')
        for linha in linhas:
            linha = linha.strip()
            if not linha:
                continue
            
            # Pattern: "Tipo Local [Titular/Suplente] [SIGLA]: Nome"
            match = re.match(r'(Gestor|Fiscal)\s+(?:Local)?\s+(?:Titular|Suplente)?\s*([A-Z]{2,})?:\s*(.+)', linha, re.IGNORECASE)
            if match:
                tipo = match.group(1).upper()
                sigla = match.group(2) or "ADM"
                nome = match.group(3).strip()
                
                if tipo == "GESTOR":
                    gestores.append((nome, sigla))
                else:
                    fiscais.append((nome, sigla))
            else:
                # Tenta format alternativo
                match = re.match(r'(Gestor|Fiscal):\s*(.+)', linha, re.IGNORECASE)
                if match:
                    tipo = match.group(1).upper()
                    nome = match.group(2).strip()
                    if tipo == "GESTOR":
                        gestores.append((nome, "ADM"))
                    else:
                        fiscais.append((nome, "ADM"))
        
        return gestores, fiscais

    def get_or_create_ativo(self, situacao):
        """Obtém ou cria ID para Ativo"""
        if not situacao:
            situacao = "Ativa"
        
        if situacao not in self.ativo_map:
            ativo_id = self.next_ativo_id
            self.ativo_map[situacao] = ativo_id
            self.sql_statements.append(
                f"INSERT INTO ativo (id, situacao) VALUES ({ativo_id}, {self.escape_sql_string(situacao)});"
            )
            self.next_ativo_id += 1
        
        return self.ativo_map[situacao]

    def get_or_create_tipo(self, tipo_name):
        """Obtém ou cria ID para Tipo"""
        if not tipo_name:
            tipo_name = "ADM"
        
        if tipo_name not in self.tipo_map:
            tipo_id = self.next_tipo_id
            self.tipo_map[tipo_name] = tipo_id
            self.sql_statements.append(
                f"INSERT INTO tipo (id, tipo_arp) VALUES ({tipo_id}, {self.escape_sql_string(tipo_name)});"
            )
            self.next_tipo_id += 1
        
        return self.tipo_map[tipo_name]

    def get_or_create_secretaria(self, sigla, nome=None):
        """Obtém ou cria ID para Secretaria"""
        if not sigla:
            sigla = "ADM"
        
        if nome is None:
            nome = sigla
        
        if sigla not in self.secretaria_map:
            sec_id = self.next_secretaria_id
            ativo_id = self.get_or_create_ativo("Ativa")
            
            self.secretaria_map[sigla] = sec_id
            self.sql_statements.append(
                f"INSERT INTO secretaria (id, nome, sigla, ativo_id) VALUES ({sec_id}, {self.escape_sql_string(nome)}, {self.escape_sql_string(sigla)}, {ativo_id});"
            )
            self.next_secretaria_id += 1
        
        return self.secretaria_map[sigla]

    def get_or_create_servidor(self, nome, sigla_secretaria):
        """Obtém ou cria ID para Servidor"""
        if nome not in self.servidor_map:
            servidor_id = self.next_servidor_id
            sec_id = self.get_or_create_secretaria(sigla_secretaria)
            ativo_id = self.get_or_create_ativo("Ativa")
            
            self.servidor_map[nome] = servidor_id
            self.sql_statements.append(
                f"INSERT INTO servidor (id, nome, cargo, matricula, email, telefone, secretaria_id, ativo_id) "
                f"VALUES ({servidor_id}, {self.escape_sql_string(nome)}, 'Servidor', 0, 'nao.informado@example.com', 'N/A', {sec_id}, {ativo_id});"
            )
            self.next_servidor_id += 1
        
        return self.servidor_map[nome]

    def get_or_create_funcao(self, funcao_name):
        """Obtém ou cria ID para FuncaoEquipe"""
        if funcao_name not in self.funcao_map:
            funcao_id = self.next_funcao_id
            ativo_id = self.get_or_create_ativo("Ativa")
            
            self.funcao_map[funcao_name] = funcao_id
            self.sql_statements.append(
                f"INSERT INTO funcao_equipe (id, nome, ativo_id) VALUES ({funcao_id}, {self.escape_sql_string(funcao_name)}, {ativo_id});"
            )
            self.next_funcao_id += 1
        
        return self.funcao_map[funcao_name]

    def generate_sql(self):
        """Gera os statements SQL"""
        wb = self.load_workbook()
        ws = wb.active
        
        # Encontrar cabeçalho
        header_row = None
        for row in range(1, min(11, ws.max_row + 1)):
            cell = ws.cell(row, 1)
            if cell.value and str(cell.value).strip().upper() in ["Nº", "N°"]:
                header_row = row
                break
        
        if not header_row:
            print("Erro: Cabeçalho não encontrado")
            sys.exit(1)
        
        # Mapear colunas
        columns = {}
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(header_row, col)
            if cell.value:
                header = str(cell.value).strip().lower()
                columns[header] = col
        
        # Adicionar cabeçalho SQL
        self.sql_statements.insert(0, "-- ===================================")
        self.sql_statements.insert(1, "-- Script de Importação de ARP")
        self.sql_statements.insert(2, f"-- Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        self.sql_statements.insert(3, "-- ===================================\n")
        self.sql_statements.insert(4, "-- IMPORTANTE: Execute este script em uma transação!")
        self.sql_statements.insert(5, "-- BEGIN; ... COMMIT; ou SET AUTOCOMMIT=1;\n")
        self.sql_statements.insert(6, "SET FOREIGN_KEY_CHECKS=0;\n")
        
        # Processar dados
        for row in range(header_row + 1, ws.max_row + 1):
            # Obter valores
            numero = ws.cell(row, columns.get('nº', columns.get('n°', 1))).value
            
            if not numero:
                continue
            
            try:
                numero = int(numero)
            except:
                continue
            
            ano = ws.cell(row, columns.get('ano', 2)).value
            data_inicio = ws.cell(row, columns.get('data início', 3)).value
            data_fim = ws.cell(row, columns.get('data fim', 4)).value
            situacao = ws.cell(row, columns.get('situação', 5)).value
            tipo = ws.cell(row, columns.get('tipo', 6)).value
            objeto = ws.cell(row, columns.get('objeto', 7)).value
            portaria_num = ws.cell(row, columns.get('nº', 8)).value
            portaria_data = ws.cell(row, columns.get('data', 9)).value
            gestor_text = ws.cell(row, columns.get('gestor', 10)).value
            fiscal_text = ws.cell(row, columns.get('fiscais', 11)).value
            
            try:
                ano = int(ano) if ano else datetime.now().year
            except:
                ano = datetime.now().year
            
            # Criar Ativo
            ativo_id = self.get_or_create_ativo(situacao)
            
            # Criar Tipo
            tipo_id = self.get_or_create_tipo(tipo)
            
            # Criar ARP
            ata_id = self.next_ata_id
            self.sql_statements.append(
                f"INSERT INTO ata_registro_preco (id, numero, ano, data_inicio, data_fim, tipo_id, objeto, ativo_id, portaria_designacao, data_designacao) "
                f"VALUES ({ata_id}, {numero}, {ano}, {self.format_date(data_inicio)}, {self.format_date(data_fim)}, {tipo_id}, {self.escape_sql_string(objeto)}, {ativo_id}, {self.escape_sql_string(portaria_num)}, {self.format_date(portaria_data)});"
            )
            self.next_ata_id += 1
            
            # Processar gestores
            gestores, fiscais = self.parse_gestores_fiscais(gestor_text)
            
            for nome, sigla in gestores:
                servidor_id = self.get_or_create_servidor(nome, sigla)
                funcao_id = self.get_or_create_funcao("GESTOR")
                
                self.sql_statements.append(
                    f"INSERT INTO equipe_contrato (id, ata_id, servidor_id, funcao_id, ativo_id) "
                    f"VALUES ({self.next_equipe_id}, {ata_id}, {servidor_id}, {funcao_id}, {ativo_id});"
                )
                self.next_equipe_id += 1
            
            # Processar fiscais
            fiscais_list, _ = self.parse_gestores_fiscais(fiscal_text)
            
            for nome, sigla in fiscais_list:
                servidor_id = self.get_or_create_servidor(nome, sigla)
                funcao_id = self.get_or_create_funcao("FISCAL")
                
                self.sql_statements.append(
                    f"INSERT INTO equipe_contrato (id, ata_id, servidor_id, funcao_id, ativo_id) "
                    f"VALUES ({self.next_equipe_id}, {ata_id}, {servidor_id}, {funcao_id}, {ativo_id});"
                )
                self.next_equipe_id += 1
        
        # Adicionar rodapé
        self.sql_statements.append("\nSET FOREIGN_KEY_CHECKS=1;")
        self.sql_statements.append("-- COMMIT; -- Descomente após verificar")
        self.sql_statements.append(f"-- Total de registros de preço: {self.next_ata_id - 1}")
        self.sql_statements.append(f"-- Total de servidores: {self.next_servidor_id - 1}")
        self.sql_statements.append(f"-- Total de equipes: {self.next_equipe_id - 1}")

    def save_sql(self, output_file):
        """Salva o SQL em arquivo"""
        with open(output_file, 'w', encoding='utf-8') as f:
            for statement in self.sql_statements:
                f.write(statement + '\n')
        
        print(f"✓ Script SQL gerado com sucesso: {output_file}")
        print(f"\nResumo:")
        print(f"  - Registros de Preço: {self.next_ata_id - 1}")
        print(f"  - Tipos criados: {self.next_tipo_id - 1}")
        print(f"  - Ativos criados: {self.next_ativo_id - 1}")
        print(f"  - Secretarias criadas: {self.next_secretaria_id - 1}")
        print(f"  - Servidores criados: {self.next_servidor_id - 1}")
        print(f"  - Funções criadas: {self.next_funcao_id - 1}")
        print(f"  - Equipes criadas: {self.next_equipe_id - 1}")

def main():
    if len(sys.argv) < 2:
        print("Uso: python generate_sql.py <arquivo_ods> [arquivo_sql_saida]")
        print("Exemplo: python generate_sql.py dados-ativos.ods import_arp.sql")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else "import_arp.sql"
    
    generator = OdsToSqlGenerator(input_file)
    generator.generate_sql()
    generator.save_sql(output_file)

if __name__ == "__main__":
    main()
