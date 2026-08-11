#!/usr/bin/env python3
"""
Script para gerar um arquivo Excel de exemplo para importação de ARP.
Requer: pip install openpyxl
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def criar_arquivo_exemplo():
    wb = Workbook()
    ws = wb.active
    ws.title = "ARP"
    
    # Cabeçalho
    headers = [
        "Nº", "Ano", "Data Início", "Data Fim", "Situação", "Tipo", "Objeto",
        "Nº", "Data",  # Portaria de Designação
        "Gestor", "Fiscais"
    ]
    
    # Estilo do cabeçalho
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Dados de exemplo 1
    dados_exemplo_1 = [
        57,  # Nº
        2022,  # Ano
        "09/09/2022",  # Data Início
        "05/09/2023",  # Data Fim
        "Desativada",  # Situação
        "ADM",  # Tipo
        "Tem como objeto, visando eventual e futura aquisição de Relógio Ponto Biométrico; Cadastrador biométrico; Crachá de aproximação e Bobina térmica. Processo: 2876/2026 Pregão Eletrônico 117/2025 realizado, alguns itens não adjudicados ARP 12/2026 Demais itens não adjudicados no Pregão 117/2025 constam no Pregão 047/2026 em andamento",  # Objeto
        "672/2023",  # Portaria Nº
        "14/04/2023",  # Data Designação
        """Gestor Local Titular ADM: Henrique Bernardes Aguiar
Gestor Local Suplente ADM: Elenice Rojas da S. Lopes
Gestor Local Titular SMED: Mariana de Lima Ferreira
Gestor Local Suplente SMED: Roselma Costa
Gestor Local Titular SMS: Maria Elisa Menezes Silva
Gestor Local Suplente SMS: Elisabeth Carvalho Gevehr""",  # Gestores
        """Fiscal Local ADM: Gislaine Roberto Rodrigues
Fiscal Local ADM: Débora Medeiros Tomaz
Fiscal Local ADM: Lucas Santos de Lima
Fiscal Local ADM: Giovani Costa André
Fiscal Local ADM: Alexandre de Lima Pereira
Fiscal Local SMED: Angelita Simas Valentim Claro
Fiscal Local SMED: Julio Carlos Rosa da Silva Neto
Fiscal Local SMED: Clayton Platen da Silva
Fiscal Local SMS: Marcia Silvestre de Oliveira
Fiscal Local SMS: Jailson Rocha
Fiscal Local SMS: jordana Groth Mingurre
Fiscal: Jordano Smorlarck Dos Santos
Fiscal: Evandro da Silva Marques"""  # Fiscais
    ]
    
    # Inserir dados exemplo 1
    for col, valor in enumerate(dados_exemplo_1, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = valor
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Dados de exemplo 2
    dados_exemplo_2 = [
        67,  # Nº
        2022,  # Ano
        "13/10/2022",  # Data Início
        "12/10/2023",  # Data Fim
        "Desativada",  # Situação
        "ADM",  # Tipo
        "Registro de preço visando eventual e futura aquisição de Bandeiras Oficiais do Brasil, Rio Grande do Sul e Imbé. ETP em fase de elaboração",  # Objeto
        "2371/2022",  # Portaria Nº
        "19/10/2022",  # Data Designação
        """Gestor Local Titular ADM: Henrique Bernardes Aguiar
Gestor Local Suplente ADM: Patrícia da Silva Dimer
Gestor Local Titular SMEC: Mariana de Lima Ferreira
Gestor Local Suplente SMEC: Roselma Costa""",  # Gestores
        """Fiscal Local ADM: Gislaine Roberto Rodrigues
Fiscal Local ADM: Débora Medeiros Tomaz
Fiscal Local ADM: Lucas Santos de Lima
Fiscal Local ADM: Giovani Costa André
Fiscal Local SMED: Angelita Simas Valentim Claro
Fiscal Local SMED: Jader Teixeira Sesterheim
Fiscal Local SMED: Clayton Platen da Silva"""  # Fiscais
    ]
    
    # Inserir dados exemplo 2
    for col, valor in enumerate(dados_exemplo_2, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = valor
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Dados de exemplo 3
    dados_exemplo_3 = [
        70,  # Nº
        2022,  # Ano
        "01/11/2022",  # Data Início
        "31/10/2023",  # Data Fim
        "Desativada",  # Situação
        "SMED",  # Tipo
        "Registro de Preço para a aquisição de mobiliário de cozinha em inox para atender demanda futura e eventual no período de 12 meses",  # Objeto
        "675/2023",  # Portaria Nº
        "14/04/2023",  # Data Designação
        """Gestor Titular: Mariana de Lima Ferreira
Gestor Suplente: Roselma Costa""",  # Gestores
        """Fiscal: Angelita Simas Valentim Claro
Fiscal: Julio Carlos Rosa da Silva Neto
Fiscal: Clayton Platen da Silva"""  # Fiscais
    ]
    
    # Inserir dados exemplo 3
    for col, valor in enumerate(dados_exemplo_3, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = valor
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 50
    ws.column_dimensions['K'].width = 50
    
    # Ajustar altura das linhas
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 100
    ws.row_dimensions[3].height = 80
    ws.row_dimensions[4].height = 60
    
    # Salvar arquivo
    arquivo_saida = "exemplo_arp.xlsx"
    wb.save(arquivo_saida)
    print(f"✓ Arquivo criado com sucesso: {arquivo_saida}")
    print(f"  Total de registros: 3")
    print(f"  Você pode usar este arquivo para testar a importação via API")

if __name__ == "__main__":
    try:
        criar_arquivo_exemplo()
    except ImportError:
        print("Erro: openpyxl não está instalado")
        print("Para instalar, execute: pip install openpyxl")
    except Exception as e:
        print(f"Erro ao criar arquivo: {e}")
